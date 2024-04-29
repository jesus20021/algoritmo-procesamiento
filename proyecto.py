import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font

ruta = "ejemplo.xlsx"
datos_excel = pd.read_excel(ruta, sheet_name=None, header=None)

concentrado = []

wb = Workbook()
hoja_activa = wb.active
hoja_activa.title = 'Fase 1 - Promedios'
hoja_activa['B2'] = 'Valor Óptimo'
hoja_activa['C2'] = 'Problema'
hoja_activa['D2'] = 'CostoProm'
hoja_activa['E2'] = 'CostoProm'
hoja_activa['F2'] = 'PeorCosto'
hoja_activa['G2'] = 'PeorCosto'
hoja_activa['H2'] = 'NFEProm'
hoja_activa['I2'] = 'NFEProm'
hoja_activa['J2'] = 'TimeProm'
hoja_activa['K2'] = 'TimeProm'

hoja_activa['B2'].font = Font(bold=True)
hoja_activa['C2'].font = Font(bold=True)
hoja_activa['D2'].font = Font(bold=True)
hoja_activa['E2'].font = Font(bold=True)
hoja_activa['F2'].font = Font(bold=True)
hoja_activa['G2'].font = Font(bold=True)
hoja_activa['H2'].font = Font(bold=True)
hoja_activa['I2'].font = Font(bold=True)
hoja_activa['J2'].font = Font(bold=True)
hoja_activa['K2'].font = Font(bold=True)
Algoritmos = True
i = 1
num_problemas = len(datos_excel.items())
for nombre_hoja, datos_hoja in datos_excel.items():
    matriz = datos_hoja.to_numpy().transpose()
    problema = matriz[0][0]
    algoritmo1 = matriz[1][0][19:]
    nfe1 = matriz[4][1:31]
    time1 = matriz[6][1:31]
    costo1 = matriz[8][1:31]
    algoritmo2 = matriz[10][0][19:]
    nfe2 = matriz[13][1:31]
    time2 = matriz[15][1:31]
    costo2 = matriz[17][1:31]

    if(Algoritmos):
        hoja_activa['D1'] = algoritmo1
        hoja_activa['F1'] = algoritmo1
        hoja_activa['H1'] = algoritmo1
        hoja_activa['J1'] = algoritmo1
        hoja_activa['E1'] = algoritmo2
        hoja_activa['G1'] = algoritmo2
        hoja_activa['I1'] = algoritmo2
        hoja_activa['K1'] = algoritmo2
        hoja_activa['D1'].font = Font(bold=True)
        hoja_activa['F1'].font = Font(bold=True)
        hoja_activa['H1'].font = Font(bold=True)
        hoja_activa['J1'].font = Font(bold=True)
        hoja_activa['E1'].font = Font(bold=True)
        hoja_activa['G1'].font = Font(bold=True)
        hoja_activa['I1'].font = Font(bold=True)
        hoja_activa['K1'].font = Font(bold=True)
        Algoritmos = False
    
    valor_optimo = input('Valor óptimo de ' + problema + ': ')
    hoja_activa['A'+str(int(i+2))] = i
    hoja_activa['A'+str(int(i+2))].font = Font(bold=True)
    hoja_activa['B'+str(int(i+2))] = float(valor_optimo)
    hoja_activa['C'+str(int(i+2))] = problema
    hoja_activa['D'+str(int(i+2))] = np.mean(costo1)
    hoja_activa['E'+str(int(i+2))] = np.mean(costo2)
    hoja_activa['F'+str(int(i+2))] = np.max(costo1)
    hoja_activa['G'+str(int(i+2))] = np.max(costo2)
    hoja_activa['H'+str(int(i+2))] = np.mean(nfe1)
    hoja_activa['I'+str(int(i+2))] = np.mean(nfe2)
    hoja_activa['J'+str(int(i+2))] = np.mean(time1)
    hoja_activa['K'+str(int(i+2))] = np.mean(time2)
    i+=1

    concentrado.append({
        'valor optimo': valor_optimo,
        'function': problema,
        'Algoritmo1': {
            'costo promedio': np.mean(costo1),
            'peor costo': np.max(costo1),
            'nfe promedio': np.mean(nfe1),
            'time promedio': np.mean(time1)
        },
        'Algoritmo2': {
            'costo promedio': np.mean(costo2),
            'peor costo': np.max(costo2),
            'nfe promedio': np.mean(nfe2),
            'time promedio': np.mean(time2)
        }
    })

for col in hoja_activa.columns:
    hoja_activa.column_dimensions[col[0].column_letter].auto_size = True
    
hoja_activa = wb.create_sheet('Fase 2 - TPS..RPS')

hoja_activa['C1'] = 'TPS'
hoja_activa['D1'] = 'TPS'
hoja_activa['E1'] = 'TPS'
hoja_activa['F1'] = 'TPS'
hoja_activa['G1'] = 'TPS'
hoja_activa['H1'] = 'TPS'
hoja_activa['J1'] = 'RPS'
hoja_activa['K1'] = 'RPS'
hoja_activa['L1'] = 'RPS'
hoja_activa['M1'] = 'RPS'
hoja_activa['N1'] = 'RPS'
hoja_activa['O1'] = 'RPS'

hoja_activa['C2'] = 'COSTO'
hoja_activa['D2'] = 'COSTO'
hoja_activa['E2'] = 'NFE'
hoja_activa['F2'] = 'NFE'
hoja_activa['G2'] = 'TIME'
hoja_activa['H2'] = 'TIME'
hoja_activa['J2'] = 'COSTO'
hoja_activa['K2'] = 'COSTO'
hoja_activa['L2'] = 'NFE'
hoja_activa['M2'] = 'NFE'
hoja_activa['N2'] = 'TIME'
hoja_activa['O2'] = 'TIME'

hoja_activa['B3'] = 'FUNCTION'
hoja_activa['C3'] = algoritmo1
hoja_activa['D3'] = algoritmo2
hoja_activa['E3'] = algoritmo1
hoja_activa['F3'] = algoritmo2
hoja_activa['G3'] = algoritmo1
hoja_activa['H3'] = algoritmo2
hoja_activa['J3'] = algoritmo1
hoja_activa['K3'] = algoritmo2
hoja_activa['L3'] = algoritmo1
hoja_activa['M3'] = algoritmo2
hoja_activa['N3'] = algoritmo1
hoja_activa['O3'] = algoritmo2
for i in range(3):
    hoja_activa['B'+str(int(i+1))].font = Font(bold=True)
    hoja_activa['C'+str(int(i+1))].font = Font(bold=True)
    hoja_activa['D'+str(int(i+1))].font = Font(bold=True)
    hoja_activa['E'+str(int(i+1))].font = Font(bold=True)
    hoja_activa['F'+str(int(i+1))].font = Font(bold=True)
    hoja_activa['G'+str(int(i+1))].font = Font(bold=True)
    hoja_activa['H'+str(int(i+1))].font = Font(bold=True)
    hoja_activa['J'+str(int(i+1))].font = Font(bold=True)
    hoja_activa['K'+str(int(i+1))].font = Font(bold=True)
    hoja_activa['L'+str(int(i+1))].font = Font(bold=True)
    hoja_activa['M'+str(int(i+1))].font = Font(bold=True)
    hoja_activa['N'+str(int(i+1))].font = Font(bold=True)
    hoja_activa['O'+str(int(i+1))].font = Font(bold=True)

RPS = {'COSTO': {'Algoritmo1': [], 'Algoritmo2': []}, 'NFE': {'Algoritmo1': [], 'Algoritmo2': []}, 'TIME': {'Algoritmo1': [], 'Algoritmo2': []}}
i = 1
for problema in concentrado:
    costo1 = ((float(problema['Algoritmo1']['costo promedio']) - float(problema['valor optimo']))/(float(problema['Algoritmo1']['peor costo']) - float(problema['valor optimo'])))
    costo2 = ((float(problema['Algoritmo2']['costo promedio']) - float(problema['valor optimo']))/(float(problema['Algoritmo2']['peor costo']) - float(problema['valor optimo'])))
    hoja_activa['A'+str(int(i+3))] = i
    hoja_activa['B'+str(int(i+3))] = problema['function']
    hoja_activa['C'+str(int(i+3))] = costo1
    hoja_activa['D'+str(int(i+3))] = costo2
    hoja_activa['E'+str(int(i+3))] = problema['Algoritmo1']['nfe promedio']
    hoja_activa['F'+str(int(i+3))] = problema['Algoritmo2']['nfe promedio']
    hoja_activa['G'+str(int(i+3))] = problema['Algoritmo1']['time promedio']
    hoja_activa['H'+str(int(i+3))] = problema['Algoritmo2']['time promedio']

    hoja_activa['A'+str(int(i+3))].font = Font(bold=True)
    hoja_activa['B'+str(int(i+3))].font = Font(bold=True)

    hoja_activa['J'+str(int(i+3))] = costo1/(np.min([costo1, costo2]))
    hoja_activa['K'+str(int(i+3))] = costo2/(np.min([costo1, costo2]))
    hoja_activa['L'+str(int(i+3))] = float(problema['Algoritmo1']['nfe promedio'])/(np.min([float(problema['Algoritmo1']['nfe promedio']), float(problema['Algoritmo2']['nfe promedio'])]))
    hoja_activa['M'+str(int(i+3))] = float(problema['Algoritmo2']['nfe promedio'])/(np.min([float(problema['Algoritmo1']['nfe promedio']), float(problema['Algoritmo2']['nfe promedio'])]))
    hoja_activa['N'+str(int(i+3))] = float(problema['Algoritmo1']['time promedio'])/(np.min([float(problema['Algoritmo1']['time promedio']), float(problema['Algoritmo2']['time promedio'])]))
    hoja_activa['O'+str(int(i+3))] = float(problema['Algoritmo2']['time promedio'])/(np.min([float(problema['Algoritmo1']['time promedio']), float(problema['Algoritmo2']['time promedio'])]))
    RPS['COSTO']['Algoritmo1'].append(costo1/(np.min([costo1, costo2])))
    RPS['COSTO']['Algoritmo2'].append(costo2/(np.min([costo1, costo2])))
    RPS['NFE']['Algoritmo1'].append(float(problema['Algoritmo1']['nfe promedio'])/(np.min([float(problema['Algoritmo1']['nfe promedio']), float(problema['Algoritmo2']['nfe promedio'])])))
    RPS['NFE']['Algoritmo2'].append(float(problema['Algoritmo2']['nfe promedio'])/(np.min([float(problema['Algoritmo1']['nfe promedio']), float(problema['Algoritmo2']['nfe promedio'])])))
    RPS['TIME']['Algoritmo1'].append(float(problema['Algoritmo1']['time promedio'])/(np.min([float(problema['Algoritmo1']['time promedio']), float(problema['Algoritmo2']['time promedio'])])))
    RPS['TIME']['Algoritmo2'].append(float(problema['Algoritmo2']['time promedio'])/(np.min([float(problema['Algoritmo1']['time promedio']), float(problema['Algoritmo2']['time promedio'])])))
    i+=1

for col in hoja_activa.columns:
    hoja_activa.column_dimensions[col[0].column_letter].auto_size = True

hoja_activa = wb.create_sheet('Fase 3 - Ps')
hoja_activa['C1'] = 'RPS'
hoja_activa['D1'] = 'RPS'
hoja_activa['E1'] = 'RPS'
hoja_activa['F1'] = 'RPS'
hoja_activa['G1'] = 'RPS'
hoja_activa['H1'] = 'RPS'

hoja_activa['C2'] = 'COSTO'
hoja_activa['D2'] = 'COSTO'
hoja_activa['E2'] = 'NFE'
hoja_activa['F2'] = 'NFE'
hoja_activa['G2'] = 'TIME'
hoja_activa['H2'] = 'TIME'

hoja_activa['C3'] = algoritmo1
hoja_activa['D3'] = algoritmo2
hoja_activa['E3'] = algoritmo1
hoja_activa['F3'] = algoritmo2
hoja_activa['G3'] = algoritmo1
hoja_activa['H3'] = algoritmo2

for i in range (3):
    hoja_activa['C'+str(int(i+1))].font = Font(bold=True)
    hoja_activa['D'+str(int(i+1))].font = Font(bold=True)
    hoja_activa['E'+str(int(i+1))].font = Font(bold=True)
    hoja_activa['F'+str(int(i+1))].font = Font(bold=True)
    hoja_activa['G'+str(int(i+1))].font = Font(bold=True)
    hoja_activa['H'+str(int(i+1))].font = Font(bold=True)

hoja_activa['A3'] = '1/np'
hoja_activa['B3'] = 'Tao'

tao = 1
for i in range(70):
    hoja_activa['A'+str(int(i+4))] = 1/num_problemas
    hoja_activa['B'+str(int(i+4))] = tao
    count = 0
    for j in RPS['COSTO']['Algoritmo1']:
        if(j<=tao):
            count+=1
    hoja_activa['C'+str(int(i+4))] = count/num_problemas
    count = 0
    for j in RPS['COSTO']['Algoritmo2']:
        if(j<=tao):
            count+=1
    hoja_activa['D'+str(int(i+4))] = count/num_problemas
    count = 0
    for j in RPS['NFE']['Algoritmo1']:
        if(j<=tao):
            count+=1
    hoja_activa['E'+str(int(i+4))] = count/num_problemas
    count = 0
    for j in RPS['NFE']['Algoritmo2']:
        if(j<=tao):
            count+=1
    hoja_activa['F'+str(int(i+4))] = count/num_problemas
    count = 0
    for j in RPS['TIME']['Algoritmo1']:
        if(j<=tao):
            count+=1
    hoja_activa['G'+str(int(i+4))] = count/num_problemas
    count = 0
    for j in RPS['TIME']['Algoritmo2']:
        if(j<=tao):
            count+=1
    hoja_activa['H'+str(int(i+4))] = count/num_problemas
    tao+=0.5

chart = LineChart()
chart.title = "COSTO"
chart.y_axis.title = ''
chart.x_axis.title = ''

values1 = Reference(hoja_activa, min_col=3, min_row=3, max_row=22)
values2 = Reference(hoja_activa, min_col=4, min_row=3, max_row=22)
chart.add_data(values1, titles_from_data=True)
chart.add_data(values2, titles_from_data=True)

hoja_activa.add_chart(chart, "J2")

chart = LineChart()
chart.title = "NFE"
chart.y_axis.title = ''
chart.x_axis.title = ''

values1 = Reference(hoja_activa, min_col=5, min_row=6, max_row=72)
values2 = Reference(hoja_activa, min_col=5, min_row=6, max_row=72)
chart.add_data(values1, titles_from_data=True)
chart.add_data(values2, titles_from_data=True)

hoja_activa.add_chart(chart, "T2")

chart = LineChart()
chart.title = "TIME"
chart.y_axis.title = ''
chart.x_axis.title = ''

values1 = Reference(hoja_activa, min_col=7, min_row=3, max_row=14)
values2 = Reference(hoja_activa, min_col=8, min_row=3, max_row=14)
chart.add_data(values1, titles_from_data=True)
chart.add_data(values2, titles_from_data=True)

hoja_activa.add_chart(chart, "N22")

wb.save("concentrado.xlsx")
